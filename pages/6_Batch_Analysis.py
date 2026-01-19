"""Batch Analysis page - Analyze multiple articles at once."""
import streamlit as st
import sys
import time
import io
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
from openpyxl import load_workbook
from utils.storage import (
    get_clients, get_outlets, get_api_key, add_batch, update_batch_article
)
from utils.article_fetcher import fetch_article, extract_domain
from utils.analyzer import analyze_article
from utils.scoring import SCORING_MODEL, calculate_total_score

st.set_page_config(
    page_title="Batch Analysis - Coverage Scorer",
    page_icon="",
    layout="wide"
)

# Import and inject custom styles
from utils.styles import (
    inject_custom_css, render_sidebar, render_score_circle,
    get_score_color, get_score_grade
)

inject_custom_css()
render_sidebar()

# Cost estimate per article (based on Claude API usage)
COST_PER_ARTICLE_LOW = 0.03
COST_PER_ARTICLE_HIGH = 0.05
DELAY_BETWEEN_REQUESTS = 1.5  # seconds


def extract_excel_with_hyperlinks(file) -> tuple[pd.DataFrame, dict, dict]:
    """
    Read an Excel file and extract both cell values and hyperlinks.

    Returns:
        - df: DataFrame with cell values
        - hyperlinks: Dict mapping (row_idx, column_name) to hyperlink URL
        - columns_with_hyperlinks: Dict mapping column name to count of hyperlinks
    """
    # Load workbook with openpyxl
    wb = load_workbook(file, data_only=False)
    ws = wb.active

    # Get headers from first row
    headers = []
    for cell in ws[1]:
        headers.append(cell.value if cell.value else f"Column_{cell.column}")

    # Extract data and hyperlinks
    data = []
    hyperlinks = {}
    columns_with_hyperlinks = {h: 0 for h in headers}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False)):
        row_data = {}
        for col_idx, cell in enumerate(row):
            col_name = headers[col_idx]
            row_data[col_name] = cell.value

            # Check for hyperlink
            if cell.hyperlink and cell.hyperlink.target:
                hyperlinks[(row_idx, col_name)] = cell.hyperlink.target
                columns_with_hyperlinks[col_name] += 1

        data.append(row_data)

    df = pd.DataFrame(data)

    # Filter to only columns that actually have hyperlinks
    columns_with_hyperlinks = {k: v for k, v in columns_with_hyperlinks.items() if v > 0}

    return df, hyperlinks, columns_with_hyperlinks


def get_url_from_cell(row_idx: int, col_name: str, cell_value, hyperlinks: dict) -> str:
    """
    Get URL from either a hyperlink or plain text cell value.
    """
    # Check if there's a hyperlink for this cell
    hyperlink_url = hyperlinks.get((row_idx, col_name))
    if hyperlink_url:
        return str(hyperlink_url).strip()

    # Fall back to cell value
    if pd.isna(cell_value) or str(cell_value).strip() == "":
        return ""
    return str(cell_value).strip()


def get_key_messages_found(analysis: dict) -> list:
    """Extract the key messages that were found in the analysis."""
    messages_found = []

    tier_2 = analysis.get("tier_2_messaging", {})

    for key in ["key_messages_high", "key_messages_medium", "key_messages_low"]:
        msg_data = tier_2.get(key, {})
        if isinstance(msg_data, dict):
            found = msg_data.get("messages_found", [])
            if found:
                messages_found.extend(found)

    return messages_found


def get_analysis_notes(analysis: dict) -> str:
    """Generate comprehensive analysis notes from the summary (no truncation)."""
    summary = analysis.get("summary", {})

    notes = []

    # Add full overall assessment (no truncation)
    assessment = summary.get("overall_assessment", "")
    if assessment:
        notes.append(assessment)

    # Add all strengths
    strengths = summary.get("strengths", [])
    if strengths:
        strengths_text = "; ".join(strengths)
        notes.append(f"Strengths: {strengths_text}")

    # Add all weaknesses
    weaknesses = summary.get("weaknesses", [])
    if weaknesses:
        weaknesses_text = "; ".join(weaknesses)
        notes.append(f"Areas to improve: {weaknesses_text}")

    return " | ".join(notes) if notes else "Analysis complete"


def process_single_article(api_key: str, url: str, client: dict, outlets: list, fallback_outlet: str = "") -> dict:
    """Process a single article and return results with detailed error logging."""
    result = {
        "url": url,
        "status": "pending",
        "score": None,
        "grade": None,
        "key_messages_found": [],
        "analysis_notes": "",
        "headline": "",
        "outlet": "",
        "error": None,
        "http_status": None
    }

    # Try to fetch the article
    success, content_or_error, metadata = fetch_article(url)

    # Always capture HTTP status if available
    result["http_status"] = metadata.get("http_status")

    if not success:
        result["status"] = "failed"
        # Use the specific error message from the fetcher
        result["error"] = content_or_error
        result["outlet"] = fallback_outlet or metadata.get("domain", "")

        # Build detailed error note for Analysis Notes column
        http_status = metadata.get("http_status")
        fetch_errors = metadata.get("fetch_errors", [])

        error_parts = []
        if http_status:
            error_parts.append(f"HTTP {http_status}")
        error_parts.append(content_or_error)

        # Include all fetch errors if there were retries
        if len(fetch_errors) > 1:
            error_parts.append(f"Attempts: {'; '.join(fetch_errors)}")

        result["analysis_notes"] = f"FETCH FAILED: {' | '.join(error_parts)}"
        return result

    # Content was fetched successfully - continue with analysis
    content = content_or_error

    headline = metadata.get("title", "")
    outlet_name = fallback_outlet or metadata.get("domain", "")
    author = metadata.get("author", "")

    result["headline"] = headline
    result["outlet"] = outlet_name

    # Try to match outlet
    outlet_info = None
    domain = extract_domain(url)
    for outlet in outlets:
        if outlet.get("domain", "").lower() in domain.lower() or domain.lower() in outlet.get("domain", "").lower():
            outlet_info = outlet
            break

    # Analyze the article
    analysis_result = analyze_article(
        api_key=api_key,
        article_text=content,
        headline=headline,
        outlet_name=outlet_name,
        author=author,
        client=client,
        outlet_info=outlet_info
    )

    if not analysis_result["success"]:
        result["status"] = "failed"
        error_msg = analysis_result.get("error", "Analysis failed")
        result["error"] = error_msg
        result["analysis_notes"] = f"ANALYSIS FAILED: {error_msg}"
        return result

    # Extract scores
    analysis = analysis_result["analysis"]
    scores = {}
    for tier_key in SCORING_MODEL.keys():
        if tier_key in analysis:
            scores[tier_key] = {}
            for factor_key, factor_data in analysis[tier_key].items():
                if isinstance(factor_data, dict):
                    scores[tier_key][factor_key] = factor_data.get("score", 0)

    score_result = calculate_total_score(scores)

    result["status"] = "success"
    result["score"] = score_result["total_score"]
    result["grade"] = get_score_grade(score_result["total_score"])
    result["key_messages_found"] = get_key_messages_found(analysis)
    result["analysis_notes"] = get_analysis_notes(analysis)
    result["full_analysis"] = analysis
    result["tier_scores"] = score_result["tier_scores"]

    return result


def main():
    st.markdown("""
    <h1 style="font-size: 2.5rem !important; font-weight: 700 !important; margin-bottom: 0.5rem !important;">
        Batch Analysis
    </h1>
    <p style="color: #6B7280; font-size: 1.1rem; margin-bottom: 2rem;">
        Analyze multiple articles at once from a CSV or Excel file
    </p>
    """, unsafe_allow_html=True)

    # Check prerequisites
    api_key = get_api_key()
    clients = get_clients()
    outlets = get_outlets()

    if not api_key:
        st.error("Please configure your Anthropic API key in Settings before analyzing articles.")
        st.stop()

    if not clients:
        st.warning("No client profiles found. Please add a client profile first.")
        st.stop()

    # Initialize session state
    if "batch_df" not in st.session_state:
        st.session_state.batch_df = None
    if "batch_results" not in st.session_state:
        st.session_state.batch_results = None
    if "processing_complete" not in st.session_state:
        st.session_state.processing_complete = False
    if "column_mapping" not in st.session_state:
        st.session_state.column_mapping = {}
    if "hyperlinks" not in st.session_state:
        st.session_state.hyperlinks = {}
    if "columns_with_hyperlinks" not in st.session_state:
        st.session_state.columns_with_hyperlinks = {}

    # Step 1: File Upload
    st.markdown("## Step 1: Upload Data")

    uploaded_file = st.file_uploader(
        "Upload your media tracking spreadsheet",
        type=["csv", "xlsx", "xls"],
        help="Upload a CSV or Excel file containing article URLs"
    )

    if uploaded_file:
        try:
            # Read the file
            if uploaded_file.name.endswith('.csv'):
                df = pd.read_csv(uploaded_file)
                st.session_state.hyperlinks = {}
                st.session_state.columns_with_hyperlinks = {}
            else:
                # Use openpyxl to extract hyperlinks from Excel files
                df, hyperlinks, columns_with_hyperlinks = extract_excel_with_hyperlinks(uploaded_file)
                st.session_state.hyperlinks = hyperlinks
                st.session_state.columns_with_hyperlinks = columns_with_hyperlinks

            st.session_state.batch_df = df
            st.success(f"Loaded {len(df)} rows from {uploaded_file.name}")

            # Show hyperlink detection info
            if st.session_state.columns_with_hyperlinks:
                cols_info = ", ".join([f"{col} ({count})" for col, count in st.session_state.columns_with_hyperlinks.items()])
                st.info(f"Detected embedded hyperlinks in columns: {cols_info}")

            # Show preview
            st.markdown("### Data Preview")
            st.dataframe(df.head(5), use_container_width=True)

        except Exception as e:
            st.error(f"Error reading file: {str(e)}")
            st.stop()

    # Step 2: Column Mapping
    if st.session_state.batch_df is not None:
        df = st.session_state.batch_df
        columns_with_hyperlinks = st.session_state.columns_with_hyperlinks

        # Build column options with hyperlink indicators
        column_options = ["(None)"]
        column_name_map = {"(None)": None}  # Map display name to actual column name

        for col in df.columns:
            if col in columns_with_hyperlinks:
                display_name = f"{col} - contains hyperlinks ({columns_with_hyperlinks[col]})"
            else:
                display_name = col
            column_options.append(display_name)
            column_name_map[display_name] = col

        st.markdown("<hr style='border: none; border-top: 1px solid #E5E7EB; margin: 2rem 0;'>", unsafe_allow_html=True)
        st.markdown("## Step 2: Map Columns")
        st.markdown("Tell us which columns contain the data we need")

        if columns_with_hyperlinks:
            st.markdown("""
            <div style="
                background: #EFF6FF;
                border: 1px solid #BFDBFE;
                border-radius: 12px;
                padding: 1rem 1.25rem;
                margin: 1rem 0;
            ">
                <strong>Hyperlinks detected!</strong><br>
                Some columns contain embedded hyperlinks. You can select these columns to extract URLs from the hyperlinks rather than the visible text.
            </div>
            """, unsafe_allow_html=True)

        col1, col2 = st.columns(2)

        with col1:
            url_column_display = st.selectbox(
                "Article URL Column *",
                options=column_options,
                index=0,
                help="Select the column containing article URLs. If URLs are embedded as hyperlinks, select the column with hyperlinks."
            )
            url_column = column_name_map[url_column_display]

            outlet_column_display = st.selectbox(
                "Outlet Name Column (Optional)",
                options=column_options,
                index=0,
                help="Select the column containing outlet names (used as fallback if URL fetch fails)"
            )
            outlet_column = column_name_map[outlet_column_display]

            headline_column_display = st.selectbox(
                "Headline Column (Optional)",
                options=column_options,
                index=0,
                help="Select the column containing article headlines (used for manual entry of failed articles)"
            )
            headline_column = column_name_map[headline_column_display]

        with col2:
            # Client selection
            client_options = {c["name"]: c for c in clients}
            selected_client_name = st.selectbox(
                "Client Profile to Score Against *",
                options=list(client_options.keys()),
                help="Choose the client to score all articles against"
            )
            selected_client = client_options[selected_client_name]

        # Batch name input
        st.markdown("<hr style='border: none; border-top: 1px solid #E5E7EB; margin: 2rem 0;'>", unsafe_allow_html=True)
        st.markdown("### Batch Identification")
        batch_name = st.text_input(
            "Batch Name *",
            placeholder="e.g., Coinbase Jan 2026 Batch",
            help="Give this batch a name to identify it later in History"
        )

        if not batch_name:
            st.warning("Please enter a batch name to continue.")

        # Store mapping
        st.session_state.column_mapping = {
            "url": url_column,
            "outlet": outlet_column,
            "headline": headline_column,
            "client": selected_client,
            "batch_name": batch_name
        }

        # Validation
        if url_column is None:
            st.warning("Please select the column containing article URLs to continue.")
            st.stop()

        # Count valid URLs (considering hyperlinks)
        url_col = st.session_state.column_mapping["url"]
        hyperlinks = st.session_state.hyperlinks
        has_hyperlinks = url_col in st.session_state.columns_with_hyperlinks

        # Build list of extracted URLs for counting and preview
        extracted_urls = []
        for idx in range(len(df)):
            url = get_url_from_cell(idx, url_col, df.iloc[idx][url_col], hyperlinks)
            if url:
                extracted_urls.append(url)

        num_articles = len(extracted_urls)

        if num_articles == 0:
            st.error("No valid URLs found in the selected column.")
            st.stop()

        # URL Preview Section
        st.markdown("<hr style='border: none; border-top: 1px solid #E5E7EB; margin: 2rem 0;'>", unsafe_allow_html=True)
        st.markdown("### URL Preview")
        if has_hyperlinks:
            st.markdown("*URLs extracted from embedded hyperlinks:*")
        else:
            st.markdown("*URLs extracted from cell values:*")

        # Show first 5 URLs
        preview_urls = extracted_urls[:5]
        for i, url in enumerate(preview_urls, 1):
            st.markdown(f"{i}. `{url[:80]}{'...' if len(url) > 80 else ''}`")

        if len(extracted_urls) > 5:
            st.markdown(f"*...and {len(extracted_urls) - 5} more URLs*")

        # Step 3: Cost Estimate
        st.markdown("<hr style='border: none; border-top: 1px solid #E5E7EB; margin: 2rem 0;'>", unsafe_allow_html=True)
        st.markdown("## Step 3: Review & Start")

        # Cost estimate
        cost_low = num_articles * COST_PER_ARTICLE_LOW
        cost_high = num_articles * COST_PER_ARTICLE_HIGH
        estimated_time = num_articles * (DELAY_BETWEEN_REQUESTS + 5)  # 5 seconds average per analysis

        st.markdown(f"""
        <div style="
            background: #ECFDF5;
            border: 1px solid #86EFAC;
            border-radius: 12px;
            padding: 1rem 1.25rem;
            margin: 1rem 0;
        ">
            <strong>Estimated API Cost:</strong> ${cost_low:.2f} - ${cost_high:.2f}<br>
            <small>Based on {num_articles} articles at ~$0.03-0.05 per article</small><br><br>
            <strong>Estimated Time:</strong> ~{int(estimated_time / 60)} minutes<br>
            <small>Processing time varies based on article length and API response time</small>
        </div>
        """, unsafe_allow_html=True)

        st.markdown(f"""
        <div style="
            background: #EFF6FF;
            border: 1px solid #BFDBFE;
            border-radius: 12px;
            padding: 1rem 1.25rem;
            margin: 1rem 0;
        ">
            <strong>What will be analyzed:</strong><br>
            - <strong>{num_articles}</strong> articles with valid URLs<br>
            - Scoring against client: <strong>{selected_client_name}</strong><br>
            - Key messages tracked: <strong>{len(selected_client.get('key_messages', []))}</strong>
        </div>
        """, unsafe_allow_html=True)

        # Diagnostic Test Fetch section
        st.markdown("<hr style='border: none; border-top: 1px solid #E5E7EB; margin: 2rem 0;'>", unsafe_allow_html=True)
        st.markdown("### Diagnostic: Test Fetch")
        st.markdown("Test the first 3 URLs to verify the fetcher is working before running full analysis.")

        test_col1, test_col2 = st.columns([1, 3])
        with test_col1:
            test_fetch_btn = st.button("Test Fetch (First 3 URLs)")

        if test_fetch_btn:
            test_urls = extracted_urls[:3]
            st.markdown("<hr style='border: none; border-top: 1px solid #E5E7EB; margin: 1rem 0;'>", unsafe_allow_html=True)
            st.markdown("#### Test Results")

            for i, test_url in enumerate(test_urls, 1):
                with st.spinner(f"Testing URL {i}/{len(test_urls)}..."):
                    success, content_or_error, metadata = fetch_article(test_url)

                    domain = metadata.get("domain", "Unknown")
                    http_status = metadata.get("http_status", "N/A")

                    if success:
                        word_count = len(content_or_error.split())
                        title = metadata.get("title", "No title")[:60]
                        st.success(f"""
                        **URL {i}: {domain}**
                        - Status: HTTP {http_status}
                        - Result: SUCCESS
                        - Words: {word_count}
                        - Title: {title}
                        """)
                    else:
                        fetch_errors = metadata.get("fetch_errors", [])
                        error_detail = content_or_error

                        st.error(f"""
                        **URL {i}: {domain}**
                        - Status: HTTP {http_status if http_status else 'N/A'}
                        - Result: FAILED
                        - Error: {error_detail}
                        """)
                        if fetch_errors and len(fetch_errors) > 1:
                            with st.expander("Show all attempt errors"):
                                for err in fetch_errors:
                                    st.markdown(f"- {err}")

                    # Small delay between test fetches
                    if i < len(test_urls):
                        time.sleep(1)

            st.markdown("<hr style='border: none; border-top: 1px solid #E5E7EB; margin: 1rem 0;'>", unsafe_allow_html=True)

        # Start processing button
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            # Disable button if batch name is missing
            batch_name_valid = bool(st.session_state.column_mapping.get("batch_name", "").strip())
            start_btn = st.button(
                "Start Batch Analysis",
                type="primary",
                use_container_width=True,
                disabled=not batch_name_valid
            )

        if start_btn:
            st.session_state.batch_results = []
            st.session_state.processing_complete = False

            # Get data
            url_col = st.session_state.column_mapping["url"]
            outlet_col = st.session_state.column_mapping["outlet"]
            headline_col = st.session_state.column_mapping.get("headline")
            client = st.session_state.column_mapping["client"]
            batch_name = st.session_state.column_mapping.get("batch_name", "").strip()
            hyperlinks = st.session_state.hyperlinks

            # Progress tracking
            progress_bar = st.progress(0)
            status_text = st.empty()
            results_container = st.container()

            results = []
            successful = 0
            failed = 0

            for idx, row in df.iterrows():
                # Extract URL using hyperlink if available, otherwise use cell value
                url = get_url_from_cell(idx, url_col, row.get(url_col, ""), hyperlinks)

                # Get spreadsheet data for fallback
                fallback_outlet = str(row.get(outlet_col, "")) if outlet_col else ""
                spreadsheet_headline = str(row.get(headline_col, "")) if headline_col else ""

                # Skip empty URLs
                if not url:
                    results.append({
                        "url": "",
                        "status": "skipped",
                        "score": None,
                        "grade": None,
                        "key_messages_found": [],
                        "analysis_notes": "No URL provided",
                        "headline": spreadsheet_headline,
                        "outlet": fallback_outlet,
                        "spreadsheet_headline": spreadsheet_headline,
                        "spreadsheet_outlet": fallback_outlet,
                        "error": "No URL"
                    })
                    continue

                # Update status
                progress = (idx + 1) / len(df)
                progress_bar.progress(progress)
                status_text.markdown(f"**Analyzing {idx + 1} of {len(df)}...** `{url[:50]}{'...' if len(url) > 50 else ''}`")

                # Process article
                result = process_single_article(
                    api_key=api_key,
                    url=url,
                    client=client,
                    outlets=outlets,
                    fallback_outlet=fallback_outlet
                )

                # Add spreadsheet data for manual entry fallback
                result["spreadsheet_headline"] = spreadsheet_headline
                result["spreadsheet_outlet"] = fallback_outlet

                results.append(result)

                if result["status"] == "success":
                    successful += 1
                else:
                    failed += 1

                # Add delay between requests
                if idx < len(df) - 1:
                    time.sleep(DELAY_BETWEEN_REQUESTS)

            # Processing complete
            st.session_state.batch_results = results
            st.session_state.processing_complete = True

            progress_bar.progress(1.0)
            status_text.empty()

            # Calculate average score for successful articles
            scores = [r["score"] for r in results if r["score"] is not None]
            avg_score = sum(scores) / len(scores) if scores else 0

            # Build article results for batch storage
            article_results = []
            for result in results:
                if result["status"] == "success":
                    article_results.append({
                        "headline": result.get("headline", ""),
                        "outlet": result.get("outlet", ""),
                        "author": result.get("author", ""),
                        "url": result["url"],
                        "total_score": result["score"],
                        "grade": result.get("grade", ""),
                        "tier_scores": result.get("tier_scores", {}),
                        "detailed_scores": result.get("full_analysis", {}),
                        "summary": result.get("full_analysis", {}).get("summary", {}),
                        "key_messages_found": result.get("key_messages_found", []),
                        "analysis_notes": result.get("analysis_notes", ""),
                        "status": "success"
                    })
                elif result["status"] == "failed":
                    article_results.append({
                        "headline": result.get("spreadsheet_headline", ""),
                        "outlet": result.get("spreadsheet_outlet") or result.get("outlet", "Unknown"),
                        "author": "",
                        "url": result["url"],
                        "total_score": None,
                        "grade": None,
                        "error": result.get("error", "Unknown error"),
                        "http_status": result.get("http_status"),
                        "status": "failed"
                    })

            # Create batch summary with all article results embedded
            batch_summary = {
                "batch_name": batch_name,
                "client": client["name"],
                "article_count": len(results),
                "successful_count": successful,
                "failed_count": failed,
                "skipped_count": sum(1 for r in results if r["status"] == "skipped"),
                "avg_score": round(avg_score, 1),
                "articles": article_results
            }
            batch_id = add_batch(batch_summary)
            st.session_state.current_batch_id = batch_id
            st.session_state.current_batch_name = batch_name

            st.success(f"Batch analysis complete! {successful} successful, {failed} failed. Saved to History.")

    # Step 4: Results Display and Export
    if st.session_state.processing_complete and st.session_state.batch_results:
        results = st.session_state.batch_results
        df = st.session_state.batch_df

        st.markdown("<hr style='border: none; border-top: 1px solid #E5E7EB; margin: 2rem 0;'>", unsafe_allow_html=True)
        st.markdown("## Results")

        # Summary stats
        successful = sum(1 for r in results if r["status"] == "success")
        failed = sum(1 for r in results if r["status"] == "failed")
        skipped = sum(1 for r in results if r["status"] == "skipped")

        scores = [r["score"] for r in results if r["score"] is not None]
        avg_score = sum(scores) / len(scores) if scores else 0

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Successful", successful)
        with col2:
            st.metric("Failed", failed)
        with col3:
            st.metric("Skipped", skipped)
        with col4:
            st.metric("Avg Score", f"{avg_score:.1f}")

        # Build results dataframe
        url_col = st.session_state.column_mapping["url"]
        outlet_col = st.session_state.column_mapping["outlet"]

        # Create output dataframe with original data + new columns
        output_df = df.copy()

        # Add new columns
        output_df["Score"] = [r["score"] for r in results]
        output_df["Grade"] = [r["grade"] for r in results]
        output_df["Key Messages Found"] = [", ".join(r["key_messages_found"]) if r["key_messages_found"] else "" for r in results]
        output_df["Analysis Notes"] = [r["analysis_notes"] for r in results]
        output_df["Analysis Status"] = [r["status"].title() for r in results]

        # Show results table
        st.markdown("### Scored Results")

        # Show just the key columns for display
        display_cols = [url_col, "Score", "Grade", "Key Messages Found", "Analysis Notes", "Analysis Status"]
        if outlet_col:
            display_cols.insert(1, outlet_col)

        # Filter to only columns that exist
        display_cols = [c for c in display_cols if c in output_df.columns]

        st.dataframe(
            output_df[display_cols],
            use_container_width=True,
            height=400
        )

        # Export options
        st.markdown("<hr style='border: none; border-top: 1px solid #E5E7EB; margin: 2rem 0;'>", unsafe_allow_html=True)
        st.markdown("## Export Results")

        col1, col2 = st.columns(2)

        with col1:
            # CSV export
            csv_buffer = io.StringIO()
            output_df.to_csv(csv_buffer, index=False)
            csv_data = csv_buffer.getvalue()

            st.download_button(
                label="Download as CSV",
                data=csv_data,
                file_name=f"batch_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv",
                use_container_width=True
            )

        with col2:
            # Excel export with proper column widths and text wrapping
            excel_buffer = io.BytesIO()
            output_df.to_excel(excel_buffer, index=False, engine='openpyxl')

            # Load the workbook to apply formatting
            excel_buffer.seek(0)
            from openpyxl.styles import Alignment
            wb_export = load_workbook(excel_buffer)
            ws_export = wb_export.active

            # Apply formatting to all columns
            for col_idx, cell in enumerate(ws_export[1], 1):
                col_letter = cell.column_letter
                cell_value = cell.value

                # Set column widths based on column type
                if cell_value == "Analysis Notes":
                    ws_export.column_dimensions[col_letter].width = 100
                elif cell_value == "Key Messages Found":
                    ws_export.column_dimensions[col_letter].width = 60
                elif cell_value in ["Score", "Grade"]:
                    ws_export.column_dimensions[col_letter].width = 12
                elif cell_value == "Analysis Status":
                    ws_export.column_dimensions[col_letter].width = 18
                else:
                    ws_export.column_dimensions[col_letter].width = 30

            # Apply text wrapping to all cells
            for row in ws_export.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Save to new buffer
            excel_buffer_formatted = io.BytesIO()
            wb_export.save(excel_buffer_formatted)
            excel_data = excel_buffer_formatted.getvalue()

            st.download_button(
                label="Download as Excel",
                data=excel_data,
                file_name=f"batch_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        # Detailed results (expandable)
        st.markdown("<hr style='border: none; border-top: 1px solid #E5E7EB; margin: 2rem 0;'>", unsafe_allow_html=True)
        st.markdown("## Detailed Results")
        st.markdown("Click on any row to see the full analysis breakdown")

        for idx, result in enumerate(results):
            if result["status"] == "success" and result.get("full_analysis"):
                score = result["score"]
                grade = result["grade"]
                color = get_score_color(score)

                # Extract and clean headline
                raw_headline = result.get("headline", "") or ""
                headline = " ".join(raw_headline.split()).strip()[:60]
                if len(raw_headline) > 60:
                    headline += "..."

                # Display headline with fallback to URL
                display_text = headline if headline else result['url'][:50]

                with st.expander(f"{grade} ({score:.1f}) - {display_text}"):
                    col1, col2 = st.columns([1, 3])

                    with col1:
                        st.markdown(f"""
                        <div style="text-align: center; padding: 1rem;">
                            <div style="font-size: 3rem; font-weight: 800; color: {color};">{score:.1f}</div>
                            <div style="font-size: 1.5rem; font-weight: 600; color: {color};">{grade}</div>
                        </div>
                        """, unsafe_allow_html=True)

                    with col2:
                        st.markdown(f"**URL:** {result['url']}")
                        st.markdown(f"**Outlet:** {result['outlet']}")
                        st.markdown(f"**Headline:** {result.get('headline', 'N/A')}")

                        if result["key_messages_found"]:
                            st.markdown("**Key Messages Found:**")
                            for msg in result["key_messages_found"]:
                                st.markdown(f"- {msg}")

                        # Summary
                        summary = result.get("full_analysis", {}).get("summary", {})
                        if summary.get("overall_assessment"):
                            st.markdown(f"**Assessment:** {summary['overall_assessment']}")

                    # Tier breakdown
                    st.markdown("<hr style='border: none; border-top: 1px solid #E5E7EB; margin: 1rem 0;'>", unsafe_allow_html=True)
                    st.markdown("**Tier Scores:**")

                    tier_scores = result.get("tier_scores", {})
                    for tier_key, tier_info in tier_scores.items():
                        score_val = tier_info.get("score", 0)
                        max_val = tier_info.get("max", 0)
                        pct = (score_val / max_val * 100) if max_val > 0 else 0
                        # Get tier name with fallback
                        tier_name = tier_info.get("name", tier_key.replace("_", " ").title())

                        st.markdown(f"**{tier_name}:** {score_val:.1f} / {max_val} ({pct:.0f}%)")

            elif result["status"] == "failed":
                http_status = result.get('http_status')
                status_text = f" (HTTP {http_status})" if http_status else ""
                with st.expander(f"FAILED{status_text} - {result['url'][:50]}"):
                    st.error(f"Error: {result.get('error', 'Unknown error')}")
                    st.markdown(f"**URL:** {result['url']}")
                    if result.get('outlet'):
                        st.markdown(f"**Outlet (fallback):** {result['outlet']}")
                    if result.get('analysis_notes'):
                        st.markdown(f"**Details:** {result['analysis_notes']}")

        # Batch info display
        st.markdown("<hr style='border: none; border-top: 1px solid #E5E7EB; margin: 2rem 0;'>", unsafe_allow_html=True)
        st.markdown("### Batch Summary")

        batch_id = st.session_state.get("current_batch_id", "")
        batch_name = st.session_state.get("current_batch_name", "")

        if batch_id and batch_name:
            st.markdown(f"""
            <div style="
                background: #EFF6FF;
                border: 1px solid #BFDBFE;
                border-radius: 12px;
                padding: 1rem 1.25rem;
                margin: 1rem 0;
            ">
                <strong>Batch Name:</strong> {batch_name}<br>
                <strong>Batch ID:</strong> <code>{batch_id}</code><br><br>
                All successful analyses have been automatically saved to History.
                You can filter by batch name on the History page.
            </div>
            """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
